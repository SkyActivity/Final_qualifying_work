# Импорт стандартных библиотек Python
import csv
from datetime import date, datetime
import re
import requests
from urllib.parse import urlencode

# Импорт Django компонентов
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth import login, logout
from django.db import transaction  # Для атомарных операций с БД
from django.db.models import Avg, Count, Sum, F, Q  # Агрегатные функции и условия запросов
from django.http import JsonResponse, HttpResponse  # HTTP ответы
from django.shortcuts import render, redirect  # Рендеринг шаблонов и редиректы
from django.core.files.storage import FileSystemStorage  # Хранение загружаемых файлов

# Импорт локальных модулей приложения
from .forms import CustomUserCreationForm  # Кастомная форма регистрации
from .models import ApplicantRecord, ApplicantMapRecord, StudentRecord  # Модели данных

# Импорт сторонних библиотек для работы с данными
import openpyxl  # Чтение/запись Excel файлов
import plotly.graph_objects as go  # Графики Plotly
import plotly.express as px  # Упрощенное создание графиков
import pandas as pd  # Обработка табличных данных


def dashboard(request):
    """Главная панель управления с общей статистикой"""
    from django.db.models import Count, Avg, Sum

    chart_data = get_chart_data()

    # Подсчет статистики для dashboard
    total_students = StudentRecord.objects.count()  # Общее количество студентов
    total_specialties = StudentRecord.objects.values('specialty').distinct().count()  # Уникальные специальности
    total_regions = StudentRecord.objects.values('region').distinct().count()  # Уникальные регионы
    avg_score = StudentRecord.objects.aggregate(Avg('avg_score'))['avg_score__avg']  # Средний балл
    avg_score = round(avg_score, 1) if avg_score else None  # Округление до 1 знака после запятой

    context = {
        'chart_data': chart_data,
        'total_students': total_students,
        'total_specialties': total_specialties,
        'total_regions': total_regions,
        'avg_score': avg_score
    }

    return render(request, 'dashboard.html', context)


def build_specialty_chart(chart_type='bar', specialty_type=None):
    """
    Строит графики по специальностям с фильтрацией
    Args:
        chart_type: тип графика ('bar', 'line', 'scatter')
        specialty_type: фильтр по конкретной специальности
    Returns:
        tuple: (HTML графика, список всех специальностей)
    """
    data = get_chart_data()
    specialties = ApplicantRecord.objects.values_list('specialty', flat=True).distinct()

    # Фильтрация данных по специальности если указана
    if specialty_type and data:
        data = [record for record in data if record.specialty == specialty_type]

    if data:
        # Важная логика: берем только последний год для каждой комбинации
        # (специальность, форма обучения, уровень образования)
        # Это нужно чтобы не суммировать места по разным годам
        by_key = {}
        for record in sorted(data, key=lambda r: (r.specialty, r.learning_form, r.education_level, -(r.year or 0))):
            key = (record.specialty, record.learning_form, record.education_level)
            if key not in by_key:
                by_key[key] = record
        data = list(by_key.values())

        # Преобразование в DataFrame для Plotly
        df = pd.DataFrame([{
            'Код': record.code,
            'Специальность': record.specialty,
            'Уровень образования': record.education_level,
            'Форма обучения': record.learning_form,
            'Всего студентов': record.total_students,
            'Бюджетные места': record.budget_rf,
            'Платные места': record.paid_students,
            'Средний балл': record.average_score
        } for record in data])

        # Создание разных типов графиков
        if chart_type == 'bar':
            fig = px.bar(
                df,
                x='Специальность',
                y=['Платные места', 'Бюджетные места', 'Всего студентов'],
                color='Форма обучения',  # Разделение по цветам для разных форм обучения
                hover_data=['Уровень образования'],  # Дополнительная информация при наведении
                title="Количество студентов и места по специальностям и форме обучения"
            )
        elif chart_type == 'line':
            fig = px.line(
                df,
                x='Специальность',
                y=['Платные места', 'Бюджетные места', 'Всего студентов'],
                color='Форма обучения',
                title="Линия студентов и мест по специальностям"
            )
        elif chart_type == 'scatter':
            fig = px.scatter(
                df,
                x='Платные места',
                y='Всего студентов',
                color='Форма обучения',
                size='Средний балл',  # Размер точек зависит от среднего балла
                hover_data=['Специальность', 'Уровень образования'],
                title="Рассеяние студентов по специальностям"
            )
        else:
            fig = go.Figure()
            fig.update_layout(title='Тип графика не поддерживается', xaxis_title='Специальность',
                              yaxis_title='Значение')

        fig.update_layout(
            yaxis_title="Значение",
            xaxis_title="Специальность"
        )
    else:
        # Пустой график если нет данных
        fig = go.Figure()
        fig.update_layout(title='Нет данных для визуализации', xaxis_title='Специальность', yaxis_title='Значение')

    return fig.to_html(full_html=False), specialties


def display_chart(request):
    """Просто отображает страницу с графиками"""
    return render(request, 'chart.html')


def register(request):
    """Регистрация нового пользователя"""
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            form.save()  # Сохранение пользователя в БД
            return redirect('login')
    else:
        form = CustomUserCreationForm()
    return render(request, 'register.html', {'form': form})


def login_view(request):
    """Страница входа (функциональность в шаблоне)"""
    return render(request, 'login.html')


def logout_view(request):
    """Выход из системы"""
    logout(request)
    return redirect('login')


def display_map(request):
    """Отображение карты с данными"""
    mapData = get_map_data()
    return render(request, 'map.html', {'result': mapData})


def get_chart_data():
    """Получение всех записей ApplicantRecord для графиков"""
    records = ApplicantRecord.objects.all()
    if not records.exists():
        return None
    return records


def normalize_gender(value):
    """
    Нормализация значения пола к формату: 'M', 'F' или 'U'
    Обрабатывает различные варианты написания (рус/англ)
    """
    if value is None:
        return 'U'  # Unknown
    text = str(value).strip().lower()
    if text in ['m', 'м', 'муж', 'мужской', 'мужчина', 'male', 'man']:
        return 'M'
    if text in ['f', 'ж', 'жен', 'женский', 'женщина', 'female', 'woman']:
        return 'F'
    return 'U'


def is_blank(value):
    """Проверка на пустое значение или строку"""
    return value is None or str(value).strip() == ''


def parse_birth_date(value):
    """
    Парсинг даты рождения из разных форматов
    Поддерживает: 'YYYY-MM-DD', 'DD.MM.YYYY', 'DD/MM/YYYY'
    """
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if is_blank(value):
        return None

    text = str(value).strip()
    for fmt in ['%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y']:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def to_float(value):
    """Безопасное преобразование в float"""
    if is_blank(value):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def to_int(value):
    """Безопасное преобразование в int"""
    if is_blank(value):
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def is_xlsx_filename(filename):
    """Проверка расширения файла"""
    return bool(filename) and str(filename).lower().endswith('.xlsx')


# Ожидаемые заголовки для файлов со студентами (английские и русские версии)
EXPECTED_STUDENT_HEADERS_EN = [
    'fio',
    'gender',
    'birth_date',
    'city',
    'region',
    'math_score',
    'russian_score',
    'profile_score',
    'avg_score',
    'specialty_code',
    'specialty',
    'learning_form',
    'education_level',
    'passing_score',
]

EXPECTED_STUDENT_HEADERS_RU = [
    'фио',
    'пол',
    'дата рождения',
    'город',
    'регион',
    'математика',
    'русский',
    'профильный предмет',
    'средний балл',
    'код специальности',
    'специальность',
    'форма обучения',
    'уровень образования',
    'проходной балл',
]


def normalize_header(value):
    """Нормализация заголовков (приведение к нижнему регистру, удаление пробелов)"""
    if value is None:
        return ''
    return str(value).strip().lower()


def validate_student_headers(sheet):
    """
    Проверка заголовков файла студентов
    Возвращает: (валидны ли заголовки, строка заголовков)
    """
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        return False, []
    headers = [normalize_header(cell) for cell in header_row]
    expected_en = [normalize_header(item) for item in EXPECTED_STUDENT_HEADERS_EN]
    expected_ru = [normalize_header(item) for item in EXPECTED_STUDENT_HEADERS_RU]
    # Проверяем соответствие либо английским, либо русским заголовкам
    return headers[:len(expected_en)] == expected_en or headers[:len(expected_ru)] == expected_ru, header_row


# Ожидаемые заголовки для основного файла данных
EXPECTED_DATA_HEADERS_RU = [
    'фио',
    'пол',
    'дата рождения',
    'город',
    'область',
    'баллы егэ по математике',
    'баллы егэ по русскому языку',
    'название профильного предмета',
    'баллы профильного предмета',
    'код',
    'специальность',
    'форма обучения',
    'уровень образования',
    'проходной балл по егэ на специальность',
    'кол-во бюджетных мест',
    'кол-во платных мест',
    'год поступления',
]

# Ожидаемые заголовки для файла карты
EXPECTED_MAP_HEADERS_RU = ['город', 'регион', 'код специальности', 'количество']
EXPECTED_MAP_HEADERS_EN = ['city', 'region', 'specialty_code', 'count']


def validate_headers(sheet, expected_list):
    """Общая функция проверки заголовков"""
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        return False, []
    headers = [normalize_header(cell) for cell in header_row]
    expected = [normalize_header(item) for item in expected_list]
    # Сравниваем только нужное количество колонок (первые N)
    return headers[:len(expected)] == expected, header_row


def upload_combined_data(request):
    """
    Основная функция загрузки данных из единого Excel файла
    Обрабатывает данные студентов и агрегирует по специальностям
    """
    if request.method == 'POST' and request.FILES['file']:
        upload_folder = 'uploads'
        upload = request.FILES['file']
        if not is_xlsx_filename(upload.name):
            return JsonResponse({'status': 'error', 'message': 'Разрешены только .xlsx файлы.'}, status=400)

        fs = FileSystemStorage(location=upload_folder)
        filename = fs.save(upload.name, upload)

        try:
            wb = openpyxl.load_workbook(fs.path(filename))
            sheet = wb.active
            has_data = False

            # Проверка заголовков
            headers_ok, _ = validate_headers(sheet, EXPECTED_DATA_HEADERS_RU)
            if not headers_ok:
                expected_list = ', '.join(EXPECTED_DATA_HEADERS_RU)
                return JsonResponse({'status': 'error', 'message': f'Неверные заголовки. Ожидаются: {expected_list}'},
                                    status=400)

            # Подготовка структур данных для обработки
            student_records = []
            aggregates = {}  # Для агрегации данных по специальностям
            new_keys = set()  # Для проверки дубликатов внутри файла
            skipped_existing = 0
            skipped_in_file = 0
            skipped_applicants = 0

            # Ключевая проверка: для одной специальности должны быть одинаковые места и проходной балл
            consistency_map = {}  # key -> (passing_score, budget_seats, paid_seats, row_num)

            # Получаем существующие записи для проверки дубликатов
            existing_keys = set(
                StudentRecord.objects.values_list(
                    'fio',
                    'birth_date',
                    'city',
                    'region',
                    'math_score',
                    'russian_score',
                    'profile_subject',
                    'profile_score',
                    'specialty_code',
                    'specialty',
                    'learning_form',
                    'education_level',
                    'passing_score',
                    'year'
                )
            )

            # Обработка строк файла (начиная со 2-й строки, так как 1-я - заголовки)
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(cell is not None and str(cell).strip() != '' for cell in row):
                    continue  # Пропускаем пустые строки

                has_data = True

                # Проверка количества колонок
                if len(row) < len(EXPECTED_DATA_HEADERS_RU):
                    return JsonResponse(
                        {'status': 'error', 'message': f'Недостаточно колонок в строке {row_num}: {row}'}, status=400)

                # Извлечение данных из строки
                fio = row[0]
                gender = row[1]
                birth_date = parse_birth_date(row[2])
                city = row[3]
                region = row[4]
                math_score = to_float(row[5])
                russian_score = to_float(row[6])
                profile_subject = row[7]
                profile_score = to_float(row[8])
                specialty_code = row[9]
                specialty = row[10]
                learning_form = row[11]
                education_level = row[12]
                passing_score = to_float(row[13])
                budget_seats = to_int(row[14])
                paid_seats = to_int(row[15])
                year = to_int(row[16])

                # Обязательные поля
                if year is None:
                    return JsonResponse({'status': 'error',
                                         'message': f'Строка {row_num}: не указан год поступления. Укажите год в колонке «Год поступления».'},
                                        status=400)

                # Валидация обязательных полей
                if is_blank(fio) or is_blank(gender) or is_blank(city) or is_blank(region):
                    return JsonResponse({'status': 'error',
                                         'message': f'Строка {row_num}: некорректные данные (ФИО, пол, город, регион).'},
                                        status=400)

                normalized_gender = normalize_gender(gender)
                if normalized_gender == 'U':
                    return JsonResponse({'status': 'error', 'message': f'Строка {row_num}: некорректный пол.'},
                                        status=400)

                if birth_date is None:
                    return JsonResponse(
                        {'status': 'error', 'message': f'Строка {row_num}: некорректная дата рождения.'}, status=400)

                if is_blank(profile_subject):
                    return JsonResponse(
                        {'status': 'error', 'message': f'Строка {row_num}: не указан профильный предмет.'}, status=400)

                if None in (math_score, russian_score, profile_score, passing_score, budget_seats, paid_seats):
                    return JsonResponse({'status': 'error',
                                         'message': f'Строка {row_num}: некорректные числовые значения (баллы или места).'},
                                        status=400)

                if is_blank(specialty_code) or is_blank(specialty) or is_blank(learning_form) or is_blank(
                        education_level):
                    return JsonResponse({'status': 'error',
                                         'message': f'Некорректные данные по специальности в строке {row_num}: {row}'},
                                        status=400)

                # ВАЖНАЯ ПРОВЕРКА: согласованность данных по специальности
                # Для одной специальности (форма, уровень, год) проходной балл и места должны совпадать
                consistency_key = (
                    str(specialty_code).strip(),
                    str(specialty).strip(),
                    str(education_level).strip(),
                    str(learning_form).strip(),
                    year
                )
                if consistency_key in consistency_map:
                    prev_pass, prev_budget, prev_paid, first_row = consistency_map[consistency_key]
                    if (passing_score != prev_pass or budget_seats != prev_budget or paid_seats != prev_paid):
                        # Если значения разные - ошибка
                        return JsonResponse({
                            'status': 'error',
                            'message': (
                                f'Для одной специальности, формы обучения, уровня и года проходной балл, '
                                f'бюджетные и платные места должны совпадать во всех строках. '
                                f'Строка {first_row}: специальность «{specialty}», {education_level}, {learning_form}, год {year} — '
                                f'проходной балл {prev_pass}, бюджетных мест {prev_budget}, платных {prev_paid}. '
                                f'Строка {row_num}: те же специальность/форма/уровень/год, но проходной балл {passing_score}, '
                                f'бюджетных мест {budget_seats}, платных {paid_seats}. Исправьте файл и загрузите снова. Данные не сохранены.'
                            )
                        }, status=400)
                else:
                    consistency_map[consistency_key] = (passing_score, budget_seats, paid_seats, row_num)

                # Расчет среднего балла и возраста
                avg_score = round((math_score + russian_score + profile_score) / 3, 2)
                today = date.today()
                age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))

                # Ключ для дедупликации студентов
                dedupe_key = (
                    str(fio).strip(),
                    birth_date,
                    str(city).strip(),
                    str(region).strip(),
                    round(math_score, 2),
                    round(russian_score, 2),
                    str(profile_subject).strip(),
                    round(profile_score, 2),
                    str(specialty_code).strip(),
                    str(specialty).strip(),
                    str(learning_form).strip(),
                    str(education_level).strip(),
                    round(passing_score, 2),
                    year
                )

                # Проверка дубликатов
                if dedupe_key in existing_keys:
                    skipped_existing += 1  # Дубликат в базе
                    continue
                if dedupe_key in new_keys:
                    skipped_in_file += 1  # Дубликат внутри файла
                    continue
                new_keys.add(dedupe_key)

                # Создание объекта StudentRecord
                student_records.append(StudentRecord(
                    fio=str(fio).strip(),
                    gender=normalized_gender,
                    age=age,
                    birth_date=birth_date,
                    city=str(city).strip(),
                    region=str(region).strip(),
                    math_score=math_score,
                    russian_score=russian_score,
                    profile_subject=str(profile_subject).strip(),
                    profile_score=profile_score,
                    avg_score=avg_score,
                    specialty_code=str(specialty_code).strip(),
                    specialty=str(specialty).strip(),
                    learning_form=str(learning_form).strip(),
                    education_level=str(education_level).strip(),
                    passing_score=passing_score,
                    year=year
                ))

                # Агрегация данных по специальностям для ApplicantRecord
                key = (
                    str(specialty_code).strip(),
                    str(specialty).strip(),
                    str(education_level).strip(),
                    str(learning_form).strip(),
                    year
                )
                aggregate = aggregates.get(key)
                if not aggregate:
                    aggregate = {
                        'sum_avg': 0,  # Сумма средних баллов
                        'count': 0,  # Количество студентов
                        'budget': budget_seats,  # Бюджетные места (одно значение на специальность)
                        'paid': paid_seats  # Платные места (одно значение на специальность)
                    }
                    aggregates[key] = aggregate

                aggregate['sum_avg'] += avg_score
                aggregate['count'] += 1
                # Места не суммируем - они должны быть одинаковые у всех строк специальности

            if not has_data:
                return JsonResponse({'status': 'error', 'message': 'Файл пустой.'}, status=400)

            # АТОМАРНАЯ ОПЕРАЦИЯ: либо все сохраняется, либо ничего
            with transaction.atomic():
                # Массовое сохранение студентов
                if student_records:
                    StudentRecord.objects.bulk_create(student_records)

                # Создание агрегированных записей по специальностям
                for key, aggregate in aggregates.items():
                    code, specialty, education_level, learning_form, year = key

                    # Проверка существующей записи
                    existing_record = ApplicantRecord.objects.filter(
                        code=code,
                        specialty=specialty,
                        education_level=education_level,
                        learning_form=learning_form,
                        year=year
                    ).first()
                    if existing_record:
                        skipped_applicants += 1
                        continue

                    # Расчет итоговых значений
                    total_students = (aggregate['budget'] or 0) + (aggregate['paid'] or 0)
                    avg_score = round(aggregate['sum_avg'] / aggregate['count'], 2) if aggregate['count'] else 0

                    # Создание записи
                    ApplicantRecord.objects.create(
                        code=code,
                        specialty=specialty,
                        education_level=education_level,
                        learning_form=learning_form,
                        total_students=total_students,
                        budget_rf=aggregate['budget'],
                        paid_students=aggregate['paid'],
                        average_score=avg_score,
                        year=year
                    )

            # Формирование ответа с информацией о пропущенных записях
            skipped_total = skipped_existing + skipped_in_file
            return JsonResponse({
                'status': 'success',
                'message': (
                    'Data uploaded successfully. '
                    f'Пропущено дубликатов студентов: {skipped_total} '
                    f'(в базе: {skipped_existing}, в файле: {skipped_in_file}). '
                    f'Пропущено дубликатов направлений: {skipped_applicants}.'
                ),
                'skipped_students_total': skipped_total,
                'skipped_students_existing': skipped_existing,
                'skipped_students_in_file': skipped_in_file,
                'skipped_applicants': skipped_applicants
            })
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})

    return JsonResponse({'status': 'error', 'message': 'Invalid request.'})


def build_student_charts(specialty=None, gender=None):
    """
    Создание набора графиков по студентам с фильтрацией
    Returns:
        dict: Словарь с HTML графиками
    """
    # Фильтрация записей
    student_records = StudentRecord.objects.all()
    if specialty:
        student_records = student_records.filter(specialty=specialty)
    if gender:
        normalized_gender = normalize_gender(gender)
        if normalized_gender in ['M', 'F']:
            student_records = student_records.filter(gender=normalized_gender)

    if not student_records.exists():
        return {}

    # Преобразование QuerySet в DataFrame для удобной работы
    df = pd.DataFrame([{
        'ФИО': record.fio,
        'Пол': normalize_gender(record.gender),
        'Возраст': record.age,
        'Дата рождения': record.birth_date,
        'Город': record.city,
        'Регион': record.region,
        'Математика': record.math_score,
        'Русский': record.russian_score,
        'Профильный': record.profile_score,
        'Профильный предмет': record.profile_subject,
        'Средний балл': round(record.avg_score, 1),
        'Код специальности': record.specialty_code,
        'Специальность': record.specialty,
        'Форма обучения': record.learning_form,
        'Уровень образования': record.education_level,
        'Проходной балл': record.passing_score,
    } for record in student_records])

    charts = {}

    # 1. Распределение по возрасту (гистограмма)
    fig_age = px.histogram(
        df,
        x='Возраст',
        nbins=15,  # Количество бинов (столбцов)
        title='Распределение студентов по возрасту'
    )
    fig_age.update_layout(template='plotly_white', height=400)
    charts['age_distribution'] = fig_age.to_html(full_html=False)

    # 2. Половозрастная пирамида
    age_gender = df.groupby(['Возраст', 'Пол']).size().reset_index(name='Количество')
    ages = sorted(df['Возраст'].dropna().unique())
    # Подсчет мужчин и женщин по возрастам
    male_counts = [int(age_gender[(age_gender['Возраст'] == age) & (age_gender['Пол'] == 'M')]['Количество'].sum()) for
                   age in ages]
    female_counts = [int(age_gender[(age_gender['Возраст'] == age) & (age_gender['Пол'] == 'F')]['Количество'].sum())
                     for age in ages]

    fig_pyramid = go.Figure()
    # Мужчины слева (отрицательные значения)
    fig_pyramid.add_trace(go.Bar(
        y=ages,
        x=[-c for c in male_counts],
        name='Мужчины',
        orientation='h'  # Горизонтальная ориентация
    ))
    # Женщины справа (положительные значения)
    fig_pyramid.add_trace(go.Bar(
        y=ages,
        x=female_counts,
        name='Женщины',
        orientation='h'
    ))
    fig_pyramid.update_layout(
        title='Пирамида пола и возраста',
        xaxis_title='Количество студентов',
        yaxis_title='Возраст',
        barmode='overlay',  # Столбцы накладываются друг на друга
        template='plotly_white',
        height=450
    )
    charts['gender_pyramid'] = fig_pyramid.to_html(full_html=False)

    # 3. Круговая диаграмма по регионам
    region_counts = df['Регион'].value_counts()
    fig_region_pie = go.Figure(data=[go.Pie(
        labels=region_counts.index,
        values=region_counts.values,
        hole=0.4,  # Кольцевая диаграмма
        textinfo='none',  # Текст только при наведении
        hovertemplate='%{label}<br>Доля: %{percent}<br>Студентов: %{value}<extra></extra>'
    )])
    fig_region_pie.update_traces(
        textposition='inside',
        pull=0,  # Не выдвигать сектора
        showlegend=True
    )
    fig_region_pie.update_layout(
        title='Распределение студентов по регионам',
        template='plotly_white',
        height=400
    )
    charts['city_region_pie'] = fig_region_pie.to_html(full_html=False)

    # 4. Heatmap: города и регионы
    top_regions = df['Регион'].value_counts().head(15).index
    top_cities = df['Город'].value_counts().head(15).index
    heat_df = df[df['Регион'].isin(top_regions) & df['Город'].isin(top_cities)]

    if not heat_df.empty:
        # Создание сводной таблицы для heatmap
        pivot_geo = pd.pivot_table(
            heat_df,
            index='Регион',
            columns='Город',
            values='ФИО',
            aggfunc='count',  # Подсчет студентов
            fill_value=0
        )
        fig_geo_heat = go.Figure(data=go.Heatmap(
            z=pivot_geo.values,  # Значения ячеек
            x=pivot_geo.columns,  # Города по оси X
            y=pivot_geo.index,  # Регионы по оси Y
            colorscale='Blues'  # Цветовая шкала
        ))
        fig_geo_heat.update_layout(
            title='Heatmap: города и регионы',
            xaxis_title='Город',
            yaxis_title='Регион',
            template='plotly_white',
            height=500
        )
        charts['geo_heatmap'] = fig_geo_heat.to_html(full_html=False)

    # 5. Гистограмма баллов ЕГЭ по предметам
    # Подготовка данных: преобразуем в длинный формат (melt)
    profile_subjects = df['Профильный предмет'].fillna('Профильный')
    profile_subjects = profile_subjects.replace('', 'Профильный')

    score_melt = pd.concat([
        pd.DataFrame({
            'Предмет': 'Математика',
            'Баллы': df['Математика']
        }),
        pd.DataFrame({
            'Предмет': 'Русский',
            'Баллы': df['Русский']
        }),
        pd.DataFrame({
            'Предмет': profile_subjects,
            'Баллы': df['Профильный']
        })
    ], ignore_index=True)

    fig_scores_hist = px.histogram(
        score_melt,
        x='Баллы',
        color='Предмет',
        barmode='overlay',
        nbins=20,
        title='Гистограмма баллов ЕГЭ по предметам'
    )
    fig_scores_hist.update_layout(template='plotly_white', height=400)
    charts['subject_histogram'] = fig_scores_hist.to_html(full_html=False)

    # 6. Boxplot распределения баллов
    fig_scores_box = px.box(
        score_melt,
        x='Предмет',
        y='Баллы',
        title='Boxplot: распределение баллов ЕГЭ по предметам'
    )
    fig_scores_box.update_layout(template='plotly_white', height=400)
    charts['subject_boxplot'] = fig_scores_box.to_html(full_html=False)

    # 7. Топ-10 специальностей по проходному баллу
    passing_top = df.groupby('Специальность')['Проходной балл'].mean().sort_values(ascending=False).head(10)
    fig_passing = px.bar(
        x=passing_top.values,
        y=passing_top.index,
        orientation='h',  # Горизонтальные столбцы
        title='Топ-10 специальностей по проходному баллу'
    )
    fig_passing.update_layout(
        xaxis_title='Проходной балл',
        yaxis_title='Специальность',
        template='plotly_white',
        height=450
    )
    charts['top_passing'] = fig_passing.to_html(full_html=False)

    # 8. Scatter plot: средний балл vs проходной балл
    # Определение статуса поступления
    df['Статус поступления'] = (df['Средний балл'] >= df['Проходной балл']).map(
        {True: 'Поступил', False: 'Не поступил'}
    )

    fig_scatter = px.scatter(
        df,
        x='Средний балл',
        y='Проходной балл',
        color='Статус поступления',
        hover_data=['ФИО', 'Специальность'],
        title='Баллы ЕГЭ vs проходной балл',
        category_orders={'Статус поступления': ['Поступил', 'Не поступил']},
        color_discrete_map={'Поступил': '#2ca02c', 'Не поступил': '#d62728'}  # Зеленый и красный
    )
    if df['Статус поступления'].nunique() < 2:
        fig_scatter.update_layout(showlegend=False)
    fig_scatter.update_layout(template='plotly_white', height=450)
    charts['score_scatter'] = fig_scatter.to_html(full_html=False)

    # 9. Heatmap популярности специальностей по формам обучения
    top_specialties = df['Специальность'].value_counts().head(15).index
    form_heat_df = df[df['Специальность'].isin(top_specialties)]
    pivot_form = pd.pivot_table(
        form_heat_df,
        index='Форма обучения',
        columns='Специальность',
        values='ФИО',
        aggfunc='count',
        fill_value=0
    )
    fig_form_heat = go.Figure(data=go.Heatmap(
        z=pivot_form.values,
        x=pivot_form.columns,
        y=pivot_form.index,
        colorscale='Viridis'
    ))
    fig_form_heat.update_layout(
        title='Heatmap: популярность специальностей по формам обучения',
        xaxis_title='Специальность',
        yaxis_title='Форма обучения',
        template='plotly_white',
        height=500
    )
    charts['form_heatmap'] = fig_form_heat.to_html(full_html=False)

    return charts


def get_map_data():
    """
    Получение данных для карты (статистика по городам/регионам)
    Включает геокодирование при необходимости
    """
    # Агрегация данных по городам и регионам
    student_stats = StudentRecord.objects.values('city', 'region').annotate(
        average_score=Avg('avg_score'),
        total_applicants=Count('id'),
        # Подсчет поступивших: средний балл >= проходного
        total_accepted=Count('id', filter=Q(avg_score__gte=F('passing_score')))
    )

    if not student_stats:
        return []

    results = []

    for stat in student_stats:
        city = stat['city']
        region = stat['region']

        # Поиск существующей записи карты
        map_record = ApplicantMapRecord.objects.filter(
            city=city,
            region=region
        ).first()

        # Создание новой записи если не существует
        if not map_record:
            lat, lon = get_coordinates_and_save(city, region)
            map_record = ApplicantMapRecord.objects.create(
                region=region,
                city=city,
                count=0,
                lat=lat,
                lon=lon,
                code=''
            )

        # Обновление координат если они нулевые
        if not map_record.lat or map_record.lat == 0 or not map_record.lon or map_record.lon == 0:
            lat, lon = get_coordinates_and_save(map_record.city, map_record.region)
            map_record.lat = lat
            map_record.lon = lon
            map_record.save()

        # Формирование результата для карты
        results.append({
            'region': region,
            'city': city,
            'count': stat['total_accepted'] or 0,
            'lat': map_record.lat,
            'lon': map_record.lon,
            'averageScore': round(stat['average_score'] or 0, 1),
            'totalApplicants': stat['total_applicants'] or 0,
            'totalAccepted': stat['total_accepted'] or 0,
        })

    return results


def upload_chart_data(request):
    """Псевдоним для загрузки данных (совместимость)"""
    return upload_combined_data(request)


def upload_map_data(request):
    """Загрузка отдельных данных для карты"""
    if request.method == 'POST' and request.FILES['file']:
        upload_folder = 'uploads'
        upload = request.FILES['file']
        if not is_xlsx_filename(upload.name):
            return JsonResponse({'status': 'error', 'message': 'Разрешены только .xlsx файлы.'}, status=400)

        fs = FileSystemStorage(location=upload_folder)
        filename = fs.save(upload.name, upload)

        try:
            wb = openpyxl.load_workbook(fs.path(filename))
            sheet = wb.active
            has_data = False

            # Проверка заголовков (русские или английские)
            headers_ok_ru, _ = validate_headers(sheet, EXPECTED_MAP_HEADERS_RU)
            headers_ok_en, _ = validate_headers(sheet, EXPECTED_MAP_HEADERS_EN)
            if not (headers_ok_ru or headers_ok_en):
                expected_list = ', '.join(EXPECTED_MAP_HEADERS_RU)
                return JsonResponse({'status': 'error', 'message': f'Неверные заголовки. Ожидаются: {expected_list}'},
                                    status=400)

            # Обработка строк
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not any(cell is not None and str(cell).strip() != '' for cell in row):
                    continue
                has_data = True

                if len(row) < 4:
                    print({'status': 'error', 'message': f'Row {row} has insufficient columns.'})
                    return JsonResponse({'status': 'error', 'message': f'Row {row} has insufficient columns.'},
                                        status=400)

                city = row[0]
                region = row[1]
                specialty_code = row[2]
                count = to_int(row[3])
                lat = 0.0
                lon = 0.0

                # Валидация
                if is_blank(city) or is_blank(region) or is_blank(specialty_code) or count is None:
                    return JsonResponse({'status': 'error', 'message': f'Некорректные данные в строке: {row}'},
                                        status=400)

                # Проверка существования специальности
                specialty = ApplicantRecord.objects.filter(code=specialty_code).first()
                if not specialty:
                    print({'status': 'error', 'message': f'Specialty "{specialty_code}" not found.'})
                    return JsonResponse({'status': 'error', 'message': f'Specialty "{specialty_code}" not found.'})

                # Геокодирование если координаты нулевые
                if lat == 0.0 or lon == 0.0:
                    lat, lon = get_coordinates_and_save(city, region)

                # Создание записи
                ApplicantMapRecord.objects.create(
                    region=str(region).strip(),
                    city=str(city).strip(),
                    count=count,
                    lat=lat,
                    lon=lon,
                    code=str(specialty_code).strip()
                )

            if not has_data:
                print({'status': 'error', 'message': 'Файл пустой.'})
                return JsonResponse({'status': 'error', 'message': 'Файл пустой.'}, status=400)

            print({'status': 'success', 'message': 'ApplicantRecord uploaded successfully.'})
            return JsonResponse({'status': 'success', 'message': 'ApplicantRecord uploaded successfully.'})

        except Exception as e:
            print({'status': 'error', 'message': str(e)})
            return JsonResponse({'status': 'error', 'message': str(e)})

    print({'status': 'error', 'message': 'Invalid request.'})
    return JsonResponse({'status': 'error', 'message': 'Invalid request.'})


def upload_students_data(request):
    """Отдельная загрузка данных студентов (устаревшая функция)"""
    if request.method == 'POST' and request.FILES['file']:
        upload_folder = 'uploads'
        upload = request.FILES['file']
        if not is_xlsx_filename(upload.name):
            return JsonResponse({'status': 'error', 'message': 'Разрешены только .xlsx файлы.'}, status=400)

        fs = FileSystemStorage(location=upload_folder)
        filename = fs.save(upload.name, upload)

        try:
            wb = openpyxl.load_workbook(fs.path(filename))
            sheet = wb.active
            has_data = False

            # Проверка заголовков
            headers_ok, header_row = validate_student_headers(sheet)
            if not headers_ok:
                expected_list = ', '.join(EXPECTED_STUDENT_HEADERS_RU)
                return JsonResponse(
                    {
                        'status': 'error',
                        'message': f'Неверные заголовки. Ожидаются: {expected_list}'
                    },
                    status=400
                )

            # Обработка строк
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not any(cell is not None and str(cell).strip() != '' for cell in row):
                    continue
                has_data = True

                if len(row) < len(EXPECTED_STUDENT_HEADERS_EN):
                    return JsonResponse(
                        {'status': 'error', 'message': f'Недостаточно колонок в строке: {row}'},
                        status=400
                    )

                # Извлечение данных
                fio = row[0]
                gender = row[1]
                birth_date = parse_birth_date(row[2])
                city = row[3]
                region = row[4]
                math_score = to_float(row[5])
                russian_score = to_float(row[6])
                profile_score = to_float(row[7])
                avg_score = to_float(row[8])
                specialty_code = row[9]
                specialty = row[10]
                learning_form = row[11]
                education_level = row[12]
                passing_score = to_float(row[13])

                # Валидация
                if is_blank(fio) or is_blank(gender) or is_blank(city) or is_blank(region):
                    return JsonResponse({'status': 'error', 'message': f'Некорректные данные в строке: {row}'},
                                        status=400)

                normalized_gender = normalize_gender(gender)
                if normalized_gender == 'U':
                    return JsonResponse({'status': 'error', 'message': f'Некорректный пол в строке: {row}'}, status=400)

                if birth_date is None:
                    return JsonResponse({'status': 'error', 'message': f'Некорректная дата рождения в строке: {row}'},
                                        status=400)

                today = date.today()
                age = today.year - birth_date.year - ((today.month, today.day) < (birth_date.month, birth_date.day))

                if None in (math_score, russian_score, profile_score, passing_score):
                    return JsonResponse({'status': 'error', 'message': f'Некорректные баллы в строке: {row}'},
                                        status=400)

                if avg_score is None:
                    avg_score = round((math_score + russian_score + profile_score) / 3, 2)

                if is_blank(specialty_code) or is_blank(specialty) or is_blank(learning_form) or is_blank(
                        education_level):
                    return JsonResponse(
                        {'status': 'error', 'message': f'Некорректные данные по специальности в строке: {row}'},
                        status=400)

                # Создание записи
                StudentRecord.objects.create(
                    fio=str(fio).strip(),
                    gender=normalized_gender,
                    age=age,
                    birth_date=birth_date,
                    city=str(city).strip(),
                    region=str(region).strip(),
                    math_score=math_score,
                    russian_score=russian_score,
                    profile_subject='',
                    profile_score=profile_score,
                    avg_score=avg_score,
                    specialty_code=str(specialty_code).strip(),
                    specialty=str(specialty).strip(),
                    learning_form=str(learning_form).strip(),
                    education_level=str(education_level).strip(),
                    passing_score=passing_score,
                    year=date.today().year  # Текущий год по умолчанию
                )

            if not has_data:
                return JsonResponse({'status': 'error', 'message': 'Файл пустой.'}, status=400)

            return JsonResponse({'status': 'success', 'message': 'StudentRecord uploaded successfully.'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})

    return JsonResponse({'status': 'error', 'message': 'Invalid request.'})


# Локальный кэш координат для популярных городов России
# Ускоряет геокодирование и снижает нагрузку на API
LOCAL_CITY_COORDS = {
    'воронеж': (51.6608, 39.2003),
    'екатеринбург': (56.8389, 60.6057),
    'казань': (55.7961, 49.1064),
    'краснодар': (45.0355, 38.9753),
    'нижний новгород': (56.2965, 43.9361),
    'новосибирск': (55.0288, 82.9235),
    'пермь': (58.0105, 56.2502),
    'ростов-на-дону': (47.2357, 39.7015),
    'саратов': (51.5331, 46.0342),
    'уфа': (54.7351, 55.9587),
}

# Локальный кэш координат для регионов
LOCAL_REGION_COORDS = {
    'воронежская область': (51.6608, 39.2003),
    'свердловская область': (56.8389, 60.6057),
    'республика татарстан': (55.7961, 49.1064),
    'краснодарский край': (45.0355, 38.9753),
    'нижегородская область': (56.2965, 43.9361),
    'новосибирская область': (55.0288, 82.9235),
    'пермский край': (58.0105, 56.2502),
    'ростовская область': (47.2357, 39.7015),
    'саратовская область': (51.5331, 46.0342),
    'республика башкортостан': (54.7351, 55.9587),
}


def normalize_geo_key(value):
    """Нормализация географических названий для поиска в кэше"""
    if value is None:
        return ''
    text = str(value).strip().lower()
    text = text.replace('ё', 'е')  # Приведение ё к е
    text = re.sub(r'^(г\.\s*)', '', text)  # Удаление "г."
    text = re.sub(r'\s+', ' ', text)  # Удаление лишних пробелов
    return text.strip()


def get_coordinates_and_save(city, region):
    """
    Геокодирование города и региона через OpenStreetMap Nominatim API
    С кэшированием в локальном словаре
    Returns:
        tuple: (широта, долгота)
    """
    url = "https://nominatim.openstreetmap.org/search"

    def normalize_geo(value):
        """Нормализация географического названия для запроса"""
        if value is None:
            return ''
        text = str(value).strip()
        text = re.sub(r'^(г\.\s*)', '', text, flags=re.IGNORECASE)
        text = re.sub(r'\s+', ' ', text)
        # Удаление типов регионов для более точного поиска
        for suffix in ['область', 'край', 'республика', 'округ']:
            text = re.sub(rf'\b{suffix}\b', '', text, flags=re.IGNORECASE).strip()
        return text

    # Очистка входных данных
    city_clean = normalize_geo(city)
    region_clean = normalize_geo(region)

    # Проверка локального кэша
    city_key = normalize_geo_key(city_clean)
    region_key = normalize_geo_key(region_clean)
    if city_key in LOCAL_CITY_COORDS:
        return LOCAL_CITY_COORDS[city_key]
    if region_key in LOCAL_REGION_COORDS:
        return LOCAL_REGION_COORDS[region_key]

    # Варианты запросов для поиска (от более точного к менее точному)
    queries = [
        f"{city}, {region}, Россия",
        f"{city_clean}, {region_clean}, Россия",
        f"{city}, Россия",
        f"{city_clean}, Россия",
        f"{region}, Россия",
        f"{region_clean}, Россия",
    ]

    # Заголовки для запроса (обязательно для Nominatim API)
    headers = {
        'User-Agent': 'VisualData Django App (https://github.com/user/visualdata)'
    }

    # Попытка найти координаты через различные запросы
    for query in queries:
        params = {
            'q': query,
            'format': 'json',
            'limit': 1,
            'countrycodes': 'ru'  # Ограничение поиска Россией
        }

        encoded_params = urlencode(params, encoding='utf-8')
        full_url = f"{url}?{encoded_params}"

        try:
            response = requests.get(full_url, headers=headers, timeout=10)
            response.raise_for_status()
            response.encoding = "utf-8"

            data = response.json()
            if data and len(data) > 0:
                lat = float(data[0]['lat'])
                lon = float(data[0]['lon'])
                return lat, lon
        except (requests.RequestException, ValueError) as e:
            print(f"Error occurred: {e}")
            continue

    # Альтернативные параметризованные запросы
    structured_params = [
        {'city': city, 'state': region, 'country': 'Russia'},
        {'city': city_clean, 'state': region_clean, 'country': 'Russia'},
        {'city': city_clean or city, 'country': 'Russia'},
        {'state': region_clean or region, 'country': 'Russia'},
    ]

    for params in structured_params:
        params.update({'format': 'json', 'limit': 1, 'countrycodes': 'ru'})
        encoded_params = urlencode(params, encoding='utf-8')
        full_url = f"{url}?{encoded_params}"
        try:
            response = requests.get(full_url, headers=headers, timeout=10)
            response.raise_for_status()
            response.encoding = "utf-8"
            data = response.json()
            if data and len(data) > 0:
                lat = float(data[0]['lat'])
                lon = float(data[0]['lon'])
                return lat, lon
        except (requests.RequestException, ValueError) as e:
            print(f"Error occurred: {e}")
            continue

    # Если координаты не найдены, возвращаем нули
    return 0, 0


def export_map_data_csv(request):
    """Экспорт данных карты в CSV файл"""
    response = HttpResponse(content_type='text/csv')
    # Установка заголовка для скачивания файла
    response['Content-Disposition'] = 'attachment; filename="map_data.csv"'

    writer = csv.writer(response)
    writer.writerow(['Город', 'Регион', 'Средний балл', 'Количество поступивших', 'Количество абитуриентов'])

    map_data = get_map_data()
    for item in map_data:
        writer.writerow([
            item['city'],
            item['region'],
            round(item['averageScore'] or 0, 1),
            item['totalAccepted'] or 0,
            item['totalApplicants'] or 0,
        ])

    return response


def clear_database(request):
    """
    Очистка базы данных (только таблиц ApplicantRecord и StudentRecord)
    Данные карты (ApplicantMapRecord) сохраняются
    """
    if request.method != 'POST':
        return JsonResponse({'status': 'error', 'message': 'Метод не разрешён.'}, status=405)

    n_app = ApplicantRecord.objects.count()
    n_stud = StudentRecord.objects.count()
    n_map = ApplicantMapRecord.objects.count()

    # Удаление данных
    ApplicantRecord.objects.all().delete()
    StudentRecord.objects.all().delete()

    return JsonResponse({
        'status': 'success',
        'message': f'База очищена. Удалено: {n_app} направлений, {n_stud} студентов. Данные карты сохранены ({n_map} записей).',
        'deleted_applicants': n_app,
        'deleted_students': n_stud,
        'map_preserved': n_map,
    })


# === НОВЫЕ ПРЕДСТАВЛЕНИЯ ДЛЯ АНАЛИТИКИ ===

def analytics_years(request):
    """Аналитика динамики поступлений по годам"""
    if not StudentRecord.objects.exists():
        context = {
            'graph_trends': go.Figure().update_layout(title='Нет данных для визуализации').to_html(full_html=False),
            'graph_scores': go.Figure().update_layout(title='Нет данных для визуализации').to_html(full_html=False),
            'years_data': [],
            'years': []
        }
        return render(request, 'analytics/years.html', context)

    # Агрегация данных по годам
    student_years = StudentRecord.objects.values('year').annotate(
        total_students=Count('id'),
        avg_score=Avg('avg_score')
    ).order_by('year')

    applicant_years = ApplicantRecord.objects.values('year').annotate(
        budget_rf=Sum('budget_rf'),
        paid_students=Sum('paid_students')
    ).order_by('year')

    # Преобразование в словари для быстрого доступа
    student_map = {
        item['year']: item
        for item in student_years
        if item['year']
    }
    applicant_map = {
        item['year']: item
        for item in applicant_years
        if item['year']
    }

    # Показываем только годы, в которых есть студенты
    years = sorted(student_map.keys())
    total_students = [student_map.get(year, {}).get('total_students', 0) for year in years]
    avg_scores = [round(student_map.get(year, {}).get('avg_score') or 0, 1) for year in years]
    budget_rf = [applicant_map.get(year, {}).get('budget_rf', 0) or 0 for year in years]
    paid = [applicant_map.get(year, {}).get('paid_students', 0) or 0 for year in years]

    # График динамики (линейный)
    fig_trends = go.Figure()
    fig_trends.add_trace(go.Scatter(
        x=years, y=total_students, name='Всего студентов',
        mode='lines+markers', line=dict(width=3), marker=dict(size=10)
    ))
    fig_trends.add_trace(go.Scatter(
        x=years, y=budget_rf, name='Бюджетные места',
        mode='lines+markers', line=dict(width=3), marker=dict(size=10)
    ))
    fig_trends.add_trace(go.Scatter(
        x=years, y=paid, name='Платные места',
        mode='lines+markers', line=dict(width=3), marker=dict(size=10)
    ))
    fig_trends.update_layout(
        title='Динамика поступлений по годам',
        xaxis_title='Год',
        yaxis_title='Количество студентов',
        hovermode='x unified',  # Общая подсказка для всех линий
        template='plotly_white',
        height=500
    )

    # График средних баллов по годам (столбчатый)
    fig_scores = go.Figure()
    fig_scores.add_trace(go.Bar(
        x=years, y=avg_scores, name='Средний балл',
        marker_color='rgb(55, 83, 109)'
    ))
    fig_scores.update_layout(
        title='Динамика средних баллов по годам',
        xaxis_title='Год',
        yaxis_title='Средний балл',
        template='plotly_white',
        height=400
    )

    graph_trends = fig_trends.to_html(full_html=False)
    graph_scores = fig_scores.to_html(full_html=False)

    # Данные для таблицы
    years_summary = [
        {
            'year': year,
            'total_students': student_map.get(year, {}).get('total_students', 0),
            'avg_score': round(student_map.get(year, {}).get('avg_score') or 0, 1)
        }
        for year in years
    ]
    context = {
        'graph_trends': graph_trends,
        'graph_scores': graph_scores,
        'years_data': years_summary,
        'years': years
    }

    return render(request, 'analytics/years.html', context)


def analytics_directions(request):
    """Аналитика по направлениям и специальностям"""
    if not StudentRecord.objects.exists():
        empty_fig = go.Figure().update_layout(title='Нет данных для визуализации')
        return render(request, 'analytics/directions.html', {
            'graph_bar': empty_fig.to_html(full_html=False),
            'graph_pie': empty_fig.to_html(full_html=False),
            'graph_levels': empty_fig.to_html(full_html=False),
            'specialties_data': []
        })

    # Получение топ-20 специальностей по количеству студентов
    specialties_qs = StudentRecord.objects.values('specialty').annotate(
        total_students=Count('id'),
        avg_score=Avg('avg_score')
    ).order_by('-total_students')[:20]

    # Сложная логика: определение бюджетных и платных мест для каждой специальности
    # У одной специальности может быть несколько записей (разные уровни образования, формы обучения, годы)
    # Берем вариант с наибольшим количеством студентов
    applicant_map = {}
    for specialty in [item['specialty'] for item in specialties_qs]:
        # Определение наиболее популярной комбинации (уровень + форма)
        main = StudentRecord.objects.filter(specialty=specialty).values(
            'education_level', 'learning_form'
        ).annotate(c=Count('id')).order_by('-c').first()
        if main:
            # Поиск соответствующей записи в ApplicantRecord
            rec = ApplicantRecord.objects.filter(
                specialty=specialty,
                education_level=main['education_level'],
                learning_form=main['learning_form']
            ).order_by('-year').first()  # Берем последний год
            if rec:
                applicant_map[specialty] = {'budget_rf': rec.budget_rf, 'paid_students': rec.paid_students}

    # Формирование данных для отображения
    specialties_data = []
    for item in specialties_qs:
        specialty = item['specialty']
        applicant = applicant_map.get(specialty)
        specialties_data.append({
            'specialty': specialty,
            'total_students': item['total_students'] or 0,
            'avg_score': item['avg_score'],
            'budget_rf': applicant['budget_rf'] if applicant else None,
            'paid_students': applicant['paid_students'] if applicant else None
        })

    # Подготовка данных для графиков
    specialties = [item['specialty'][:30] + '...' if len(item['specialty']) > 30 else item['specialty']
                   for item in specialties_data]
    students = [item['total_students'] or 0 for item in specialties_data]

    # 1. Горизонтальная столбчатая диаграмма топ-20 специальностей
    fig_bar = go.Figure()
    fig_bar.add_trace(go.Bar(
        x=students,
        y=specialties,
        orientation='h',  # Горизонтальная ориентация
        marker=dict(
            color=students,  # Цвет зависит от количества студентов
            colorscale='Viridis',  # Цветовая шкала
            showscale=True,  # Показывать шкалу цветов
            colorbar=dict(title="Студентов")
        ),
        text=students,  # Текст на столбцах
        textposition='inside'
    ))
    fig_bar.update_layout(
        title='Топ-20 специальностей по количеству студентов',
        xaxis_title='Количество студентов',
        yaxis_title='Специальность',
        template='plotly_white',
        height=max(600, len(specialties) * 30),  # Динамическая высота
        margin=dict(l=200)  # Отступ слева для длинных названий
    )

    # 2. Круговая диаграмма по формам обучения
    forms_data = StudentRecord.objects.values('learning_form').annotate(
        total=Count('id')
    ).order_by('-total')

    forms = [item['learning_form'] for item in forms_data]
    forms_totals = [item['total'] or 0 for item in forms_data]

    fig_pie = go.Figure(data=[go.Pie(
        labels=forms,
        values=forms_totals,
        hole=0.4,  # Кольцевая диаграмма
        textinfo='label+percent',  # Показывать метку и процент
        textposition='outside'
    )])
    fig_pie.update_layout(
        title='Распределение по формам обучения',
        template='plotly_white',
        height=500
    )

    # 3. Столбчатая диаграмма по уровням образования
    levels_data = StudentRecord.objects.values('education_level').annotate(
        total=Count('id')
    ).order_by('-total')

    levels = [item['education_level'] for item in levels_data]
    levels_totals = [item['total'] or 0 for item in levels_data]

    fig_levels = go.Figure()
    fig_levels.add_trace(go.Bar(
        x=levels,
        y=levels_totals,
        marker_color='rgb(158,202,225)',
        text=levels_totals,
        textposition='outside'
    ))
    fig_levels.update_layout(
        title='Распределение по уровням образования',
        xaxis_title='Уровень образования',
        yaxis_title='Количество студентов',
        template='plotly_white',
        height=400
    )

    graph_bar = fig_bar.to_html(full_html=False)
    graph_pie = fig_pie.to_html(full_html=False)
    graph_levels = fig_levels.to_html(full_html=False)

    context = {
        'graph_bar': graph_bar,
        'graph_pie': graph_pie,
        'graph_levels': graph_levels,
        'specialties_data': specialties_data[:10],  # Только топ-10 для таблицы
    }

    return render(request, 'analytics/directions.html', context)


def analytics_regions(request):
    """Аналитика по регионам и городам"""
    # Агрегация данных по регионам
    regions_data = StudentRecord.objects.values('region').annotate(
        total_cities=Count('city', distinct=True),  # Уникальные города
        total_applicants=Count('id'),  # Все абитуриенты
        # Поступившие по баллам
        total_accepted=Count('id', filter=Q(avg_score__gte=F('passing_score'))),
        avg_score=Avg('avg_score')
    ).order_by('-total_applicants')  # Сортировка по убыванию количества абитуриентов

    regions_stats = [{
        'region': item['region'],
        'total_cities': item['total_cities'],
        'total_applicants': item['total_applicants'],
        'total_students': item['total_accepted'],
        'total_accepted': item['total_accepted'],
        'avg_score': round(item['avg_score'] or 0, 1),
    } for item in regions_data]

    # Подготовка данных для графиков
    regions = [item['region'] for item in regions_stats]
    total_applicants_list = [item['total_applicants'] for item in regions_stats]
    total_accepted_list = [item['total_accepted'] for item in regions_stats]

    # 1. Сгруппированная столбчатая диаграмма: абитуриенты vs поступившие
    fig_regions = go.Figure()
    fig_regions.add_trace(go.Bar(
        name='Всего абитуриентов',
        x=regions,
        y=total_applicants_list,
        marker_color='rgb(55, 83, 109)'  # Темно-синий
    ))
    fig_regions.add_trace(go.Bar(
        name='Поступило по баллам',
        x=regions,
        y=total_accepted_list,
        marker_color='rgb(26, 118, 255)'  # Светло-синий
    ))
    fig_regions.update_layout(
        title='Сравнение регионов',
        xaxis_title='Регион',
        yaxis_title='Количество',
        barmode='group',  # Сгруппированные столбцы
        template='plotly_white',
        height=500,
        xaxis_tickangle=-45  # Наклон подписей оси X
    )

    # 2. Столбчатая диаграмма средних баллов с цветовой шкалой
    avg_scores = [item['avg_score'] for item in regions_stats]

    fig_scores = go.Figure()
    fig_scores.add_trace(go.Bar(
        x=regions,
        y=avg_scores,
        marker=dict(
            color=avg_scores,
            colorscale='RdYlGn',  # Красно-желто-зеленая шкала (хорошие баллы - зеленые)
            showscale=True,
            colorbar=dict(title="Средний балл")
        ),
        text=avg_scores,
        textposition='outside'
    ))
    fig_scores.update_layout(
        title='Средние баллы по регионам',
        xaxis_title='Регион',
        yaxis_title='Средний балл',
        template='plotly_white',
        height=400,
        xaxis_tickangle=-45
    )

    # 3. Столбчатая диаграмма количества городов по регионам
    cities_count = [item['total_cities'] for item in regions_stats]

    fig_cities = go.Figure()
    fig_cities.add_trace(go.Bar(
        x=regions,
        y=cities_count,
        marker_color='rgb(158,202,225)',  # Голубой
        text=cities_count,
        textposition='outside'
    ))
    fig_cities.update_layout(
        title='Количество городов по регионам',
        xaxis_title='Регион',
        yaxis_title='Количество городов',
        template='plotly_white',
        height=400,
        xaxis_tickangle=-45
    )

    graph_regions = fig_regions.to_html(full_html=False)
    graph_scores = fig_scores.to_html(full_html=False)
    graph_cities = fig_cities.to_html(full_html=False)

    context = {
        'graph_regions': graph_regions,
        'graph_scores': graph_scores,
        'graph_cities': graph_cities,
        'regions_stats': regions_stats
    }

    return render(request, 'analytics/regions.html', context)


def analytics_students(request):
    """Аналитика по студентам с фильтрами"""
    # Получение параметров фильтрации из GET-запроса
    specialty_filter = request.GET.get('specialty')
    gender_filter = request.GET.get('gender')
    chart_type = request.GET.get('chart_type', 'bar')
    specialty_type = request.GET.get('specialty_type')

    # Построение графиков с фильтрами
    student_charts = build_student_charts(specialty_filter, gender_filter)

    # Получение списка всех специальностей для выпадающего списка
    specialties = StudentRecord.objects.values_list('specialty', flat=True).distinct().order_by('specialty')

    # Построение графика по специальностям
    graph_html, chart_specialties = build_specialty_chart(chart_type, specialty_type)

    return render(request, 'analytics/students.html', {
        'student_charts': student_charts,
        'specialties': specialties,
        'specialty_filter': specialty_filter,
        'gender_filter': gender_filter,
        'graph_html': graph_html,
        'chart_type': chart_type,
        'specialty_type': specialty_type,
        'chart_specialties': chart_specialties
    })